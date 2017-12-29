"""Creates functions for converting between data formats"""

from calendar import day_name
from csv import DictReader
from datetime import date, timedelta
from io import BytesIO, FileIO, StringIO, TextIOWrapper
from typing import BinaryIO, Dict, List, TextIO, Union

from openpyxl import load_workbook
from ruamel.yaml import YAML
from ruamel.yaml.comments import CommentedMap


FileIO = Union[BinaryIO, BytesIO, FileIO, StringIO, TextIO, TextIOWrapper]


def convert_csv_to_dict(
    data_csv_fp: FileIO,
    *,
    key: str = None,
) -> Dict[str, Dict[str, str]]:
    """Converts a CSV file to a dictionary of dictionaries

    This function inputs a CSV file and a key column (defaulting to the
    left-most column, if not specified) and outputs a dictionary keyed by the
    specified key column and having as values dictionaries encoding the row
    from the CSV file corresponding to the key value

    Args:
        data_csv_fp: pointer to CSV file or file-like object with columns
            headers in its first row and ready to be read from
        key: a column header from data_csv_fp, whose values should be used as
            keys in the dictionary generated

    Returns:
        A dictionary keyed by the specified key column and having as values
        dictionaries encoding the row from the CSV file corresponding to the
        key value

    """

    csv_file_reader = DictReader(data_csv_fp)
    if key is None:
        key = csv_file_reader.fieldnames[0]

    return_value = {}
    for row in csv_file_reader:
        return_value[row[key]] = row

    return return_value


def convert_csv_to_multimap(
    data_csv_fp: FileIO,
    *,
    key_column: str = None,
    values_column: str = None,
    overwrite_values: bool = False,
) -> Dict[str, Union[str, List[str]]]:
    """Converts a CSV file to a dictionary of dictionaries

    This function inputs a CSV file, a key column (defaulting to the
    left-most column, if not specified), a values column (defaulting to the
    second column from the left, if not specified), and a flag for whether
    values should be overwritten (defaulting to False), and outputs a
    dictionary keyed by the specified key column and having as values the
    entries from the specified values column, with values collected into a
    list if overwrite_values is False, and the last value found when reading
    the file otherwise

    Args:
        data_csv_fp: pointer to CSV file or file-like object with columns
            headers in its first row and ready to be read from
        key_column: a column header from data_csv_fp, whose values should be
            used as keys in the dictionary generated
        values_column: a column header from data_csv_fp, whose values should be
            used as values in the dictionary generated
        overwrite_values: determines where the last-appearing value
            corresponding to each key is returned; if False, then a list of
            values is returned for each key

    Returns:
        A dictionary keyed by the specified key column and having as values the
        entries from the specified values column, with values collected into a
        list if overwrite_values == False, and the last value found when
        reading the file otherwise

    """

    csv_file_reader = DictReader(data_csv_fp)
    if key_column is None:
        key_column = csv_file_reader.fieldnames[0]
    if values_column is None:
        values_column = csv_file_reader.fieldnames[1]

    return_value = {}
    for row in csv_file_reader:
        if overwrite_values:
            return_value[row[key_column]] = row[values_column]
        elif row[key_column] in return_value.keys():
            return_value[row[key_column]].append(row[values_column])
        else:
            return_value[row[key_column]] = [row[values_column]]

    return return_value


def convert_xlsx_to_dict(
    data_xlsx_fp: FileIO,
    *,
    key: str = None,
    worksheet: str = None,
) -> Dict[str, Dict[str, str]]:
    """Converts an XLSX file to dictionary of dictionaries

    This function inputs an XLSX file, a key column (defaulting to the
    left-most column, if not specified), and a worksheet name (defaulting to
    the first worksheet, if not specified), and outputs a dictionary keyed by
    the specified key column and having as values dictionaries encoding the row
    from the specified worksheet of the XLSX file corresponding to the key
    value

    Args:
        data_xlsx_fp: pointer to an XLSX file or file-like object with columns
            headers in its first row and ready to be read from
        key: a column header from data_xlsx_fp, whose values should be used as
            keys in the dictionary generated
        worksheet: a worksheet name from data_xlsx_fp, whose values should be
            used in the dictionary generated

    Returns:
        A dictionary keyed by the specified key column and having as values
        dictionaries encoding the row from the specified worksheet of the XLSX
        file corresponding to the key value

    """

    xlsx_file_reader = load_workbook(
        data_xlsx_fp,
        read_only=True,
        data_only=True
    )
    if worksheet is None:
        worksheet = xlsx_file_reader[0]
    if key is None:
        key = worksheet.cell(row=1, column=1)

    xlsx_worksheet_reader = xlsx_file_reader[worksheet]
    xlsx_worksheet_columns = xlsx_worksheet_reader.rows
    xlsx_worksheet_headers = [
        cell.value
        for cell in next(xlsx_worksheet_columns)
    ]
    key_column_index = xlsx_worksheet_headers.index(key)

    return_value = {}
    for i, row in enumerate(xlsx_worksheet_columns):
        new_row_to_add = {}
        for j, cell in enumerate(row):
            new_row_to_add[xlsx_worksheet_headers[j]] = row[j].value
        return_value[row[key_column_index].value] = new_row_to_add

    return return_value


def convert_xlsx_to_yaml_calendar(
    data_xlsx_fp: FileIO,
    start_date: date,
    *,
    item_delimiter: str = '|',
    relative_week_number_column: str = None,
    worksheet: str = None,
) -> str:
    """Converts an XLSX file to a YAML string representing a weekly calendar

    This function inputs an XLSX file, a start date, an item delimiter for
    decomposing Excel-file cell values into lists (defaulting to a vertical
    pipe), a key column for week numbers (defaulting to the left-most column,
    if not specified), and a worksheet name (defaulting to the first worksheet,
    if not specified), and outputs a string containing a YAML representation
    of the XLSX file as a dictionary keyed by the specified key column and
    having as values dictionaries encoding the row from the specified worksheet
    of the XLSX file corresponding to the key value

    Args:
        data_xlsx_fp: pointer to an XLSX file or file-like object with columns
            headers in its first row and ready to be read from; any column
            names in data_xlsx_fp corresponding to day names in the current
            locale, as identified by the calendar module, are treated as
            providing activities for the corresponding calendar date and will
            be ordered according to ISO 8601 in output; all other columns are
            treated as providing information about the week itself
        start_date: specifies the start date for the calendar, which is
            adjusted to the Monday of the week that the start_date appears in,
            per ISO 8601's specification that weeks run from Monday to Sunday
        item_delimiter: a string whose values will be used to split item values
            into lists
        relative_week_number_column: a column header from data_xlsx_fp, whose
            values should be used as key values in the YAML string generated;
            the values of relative_week_number_column should be integers, with
            the value one (1) representing the week that start_date appears in
        worksheet: a worksheet name from data_xlsx_fp, whose values should be
            used in the dictionary generated

    Returns:
         A string containing a YAML representation of the XLSX file as a
         dictionary keyed by the specified key column and having as values
         dictionaries encoding the row from the specified worksheet of the XLSX
         file corresponding to the key value

    """

    data_dict = convert_xlsx_to_dict(
        data_xlsx_fp,
        key=relative_week_number_column,
        worksheet=worksheet
    )

    start_date_adjusted = start_date - timedelta(days=start_date.weekday())

    weekdays_lookup_dict = {day.lower(): n for n, day in enumerate(day_name)}

    calendar_dict = CommentedMap()
    for week_number, week_data in data_dict.items():
        week_number = int(week_number)
        calendar_dict[week_number] = CommentedMap()
        for weekday in week_data:
            if (
                weekday == relative_week_number_column or
                week_data[weekday] is None
            ):
                continue
            if weekday.lower() in weekdays_lookup_dict:
                weekday_date = (
                    start_date_adjusted
                    +
                    timedelta(
                        days=7 * (int(week_number) - 1) +
                        int(weekdays_lookup_dict[weekday.lower()])
                    )
                ).strftime('%d%b%Y').upper()
                calendar_dict[week_number][weekday] = CommentedMap()
                calendar_dict[week_number][weekday]['Date'] = weekday_date
                calendar_dict[week_number][weekday]['Activities'] = (
                    week_data[weekday].split(item_delimiter)
                )

            else:
                calendar_dict[week_number][weekday] = (
                    week_data[weekday].split(item_delimiter)
                )

    yaml = YAML()
    calendar_yaml = StringIO()
    yaml.dump(data=calendar_dict, stream=calendar_yaml)
    calendar_yaml.seek(0)

    return calendar_yaml.read()


def flatten_dict(
    data_items: dict,
    key_value_separator: str = '\n',
    items_separator: str = '\n',
    *,
    suppress_keys: bool = False,
    **kwargs
) -> str:
    """Converts a dictionary to a string with specified separators

    This function converts dictionary data_items to a string, with
    key_value_separator used to separate keys from values and items_separator
    used to separate items; in addition, keyword arguments can be passed to the
    builtin function sorted

    Args:
        data_items: a dictionary whose keys and items will be treated as or
            converted to strings
        key_value_separator: used to separate keys from values in the output
            string
        items_separator: used to separate items in the output string
        suppress_keys: Boolean for determining whether to include keys in the
            returned string
        **kwargs: options passed through to the builtin function sorted

    Returns:
        A string representation of data_items with key_value_separator used to
        separate keys from values and items_separator used to separate items

    """

    return_value = items_separator
    last_record_number = len(data_items)
    for n, k in enumerate(sorted(data_items.keys(), **kwargs)):
        if not suppress_keys:
            return_value += str(k)
            return_value += key_value_separator
        return_value += str(data_items[k])
        if n < last_record_number - 1:
            return_value += items_separator

    return return_value
